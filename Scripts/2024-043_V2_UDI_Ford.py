# %%
#LLamar a las bibliotecas
import pandas as pd
import numpy as np
from pathlib import Path
import os
from openpyxl import workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

PathC = "C:/OTMX"
NameItem = "2024-043"


# %%
#Seleccionamos la base llamda malla
#Definimos la ruta del archivo
xlsx_dir = Path(PathC) / "Inputs" / NameItem

# %%
# Define & create output directory
pdf_output_dir = Path(PathC) / "Outputs" / NameItem
pdf_output_dir.mkdir(parents=True, exist_ok=True)

# %%

#Dataframe
Name01 = "Layout_UDI_Ford.xlsx"
PathFileName01 = xlsx_dir / Name01
df01 = pd.read_excel(PathFileName01)
df01.columns

# %%
#Sacamos los valores únicos de la agencias
df01["Agencia"].unique()

# %%
#Establecemos el bucle para el análisis
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook


# %%
# Puedes crear un objeto de estilo para el fondo azul y letras blancas
fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
font = Font(color="FFFFFF")


# %%
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook

for x in df01["Agencia"].unique():
    df_temp = df01[df01["Agencia"]== x]
    #Convertimos el archivo de excel
    df_temp.to_excel(pdf_output_dir / f"Archivo_UDI_Agencia_{x}.xlsx", index = False)

    #Edición del archivo de excel
    wb = openpyxl.load_workbook(pdf_output_dir / f"Archivo_UDI_Agencia_{x}.xlsx")
    sheet = wb.active
    
    ancho = 20
    # Recorre todas las columnas y establece el ancho deseado
    for columna in sheet.columns:
        sheet.column_dimensions[columna[0].column_letter].width = ancho
        #sheet.column_dimensions[columna[0].column_letter].fill = fill
 
    #Insertar Filas

    #Insertar logo Marsh

    #Guardar el archivo
    wb.save(pdf_output_dir / f"Archivo_UDI_Agencia_{x}.xlsx")



