# %%
print("Trace | importar Librerías")
#LLamar a las bibliotecas
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
import os
from openpyxl import workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

PathC = "C:/OTMX"
NameItem = "2024-109"

# %%
# Define & create output directory
pdf_output_dir = Path(PathC) / "Outputs" / NameItem
pdf_output_dir.mkdir(parents=True, exist_ok=True)

# %% [markdown]
# Catálogos

# %%
#Carga de los catálogos
print('Trace | Catalogues Files')

# %%
print('Trace | Catalogues Files | Pólizas')

#Catálogo de póliza
#Carga de catálogo general
dfcat_pol = pd.read_excel(Path(PathC) / "Catalogues" / "c_2024_109.xlsx", sheet_name="POL")
dfcat_pol['GMM'] = dfcat_pol['GMM'].str.replace(' ', '')
dfcat_pol['GMM LENTES'] = dfcat_pol['GMM LENTES'].str.replace(' ', '')

#Renomrar columna para el cruce
#dfcat_pol = dfcat_pol.rename(columns={'GMM':'Key'})
dfcat_pol['Key_Pol'] = dfcat_pol['GMM']

dfcat_pol.columns


# %%
print('Trace | Catalogues Files | Sitios')

#Catálogo de sitios
#Carga de catálogo general
dfcat_sit = pd.read_excel(Path(PathC) / "Catalogues" / "c_2024_109.xlsx", sheet_name="SITIO")
# Separar la información en una columna basada en un guion medio ("-")
dfcat_sit[['Pol', 'Sit']] = dfcat_sit['Numero MAPP'].str.split('-', expand=True)
dfcat_sit = dfcat_sit.rename(columns={'Numero MAPP':'POLIZA'})

#Convertir a tipo entero
dfcat_sit['Sit'] = dfcat_sit['Sit'].astype(int)

#dfcat_sit['dos_ultimos'] = dfcat_sit['Pol'].str[-2:]

dfcat_sit['Sit'] = dfcat_sit['Sit'].astype(str)
dfcat_sit['Key'] = dfcat_sit['Pol'] + "-" + dfcat_sit['Sit']

dfcat_sit.columns

# %%
print('Trace | Data File | ')
# Ejemplo de uso
#carpeta = "C:/OTMX/Inputs/2024-109"
carpeta = Path(PathC) / 'Inputs' / NameItem
NameFile = 'DATA_FORVIA_CLIENTE.xlsx'
print(NameFile)

# %%
df = pd.read_excel(carpeta / NameFile)
#df = df.rename(columns={'Sitio':'POLIZA'})
# Separar la información en una columna basada en un guion medio ("-")
df[['Pol', 'Sit']] = df['Sitio'].str.split('-', expand=True)
#df.columns.array


# %%
#
#df['F. Nacimiento'] = df['F. Nacimiento'].dt.strftime('%d-%m-%Y')

# %%
#Convertir a tipo entero
df['Sit'] = df['Sit'].astype(str)
df['Key'] = df['Pol'] + "-" + df['Sit']
df['Key_Pol'] = df['Pol']


#
df['No de Certificado'] = df['# Empleado']

# %% [markdown]
# Cruce de datos

# %%
# Utilizar la función BuscarV para buscar el valor de 'Edad' en df2 basado en 'ID' en df1
#df['SITIO'] = df['POLIZA'].map(dfcat_sit.set_index('POLIZA')['SITIO'])
#df.columns

# %%
df = pd.merge(df,dfcat_sit,on='Key',how='left')

# %%

df = pd.merge(df,dfcat_pol,on='Key_Pol',how='left')
df['Observaciones'] = df['Descripción del cambio']
df.columns



# %%
#Edicion de la fecha de nacimmiento 
df['F. Nacimiento'] = pd.to_datetime(df['F. Nacimiento'], format='%d/%m/%Y %I:%M:%S %p')
df['F. Nacimiento'] = df['F. Nacimiento'].dt.strftime('%d/%m/%Y')

#Edicion de la fecha de movimiento 
df['F. Alta/Baja'] = pd.to_datetime(df['F. Alta/Baja'], format='%d/%m/%Y %I:%M:%S %p')
df['F. Alta/Baja'] = df['F. Alta/Baja'].dt.strftime('%d/%m/%Y')


# %%

#Edición No certificado 
df['No de Certificado'] = df['No de Certificado'].astype(str)
df['No de Certificado'] = df['No de Certificado'].str[1:]

# %%
#Seleccionar las columnas
df_xlsx = df[['Razón social','GMM', 'GMM LENTES','SITIO','# Empleado','No de Certificado',
              'A. Paterno', 'A. Materno', 'Nombre Completo','F. Nacimiento','Género', 'Parentesco','F. Alta/Baja','Movimiento','Observaciones']]

#Renombrar las columnas

df_xlsx = df_xlsx.rename(columns={'Razón social':'Contratante',
                                  'GMM':'POLIZA GMM',
                                  'GMM LENTES':'POLIZA LENTES',
                                  'SITIO':'SITIO',
                                  '# Empleado':'No Empleado',
                                  'No de Certificado':'No Certificado',
                                  'A. Paterno':'Apellido Paterno',
                                  'A. Materno':'Apellido Materno',
                                  'Nombre Completo':'Nombre(s)',
                                  'F. Nacimiento':'Fecha de Nacimiento',
                                  'Género':'Género',
                                  'Parentesco':'Parentesco',
                                  'F. Alta/Baja':'Fecha Movimiento',
                                  'Movimiento':'Tipo Movimiento',
                                  'Observaciones':'Observaciones'                                                         
                                  })

df_xlsx.columns


# %%
print('Trace | Guardar archivo excel')

from datetime import date
fecha_actual = date.today().strftime("%d-%m-%Y")
nombre_archivo = f"Base_Forvia_{fecha_actual}.xlsx"


#Guardar los archivos de excel
df_xlsx.to_excel(Path(PathC) / 'Outputs' / NameItem / nombre_archivo, index=False)

#Edición del archivo de excel
wb = openpyxl.load_workbook( Path(PathC) / 'Outputs' / NameItem / nombre_archivo)
sheet = wb.active
ancho = 20
# Recorre todas las columnas y establece el ancho deseado
for columna in sheet.columns:
    sheet.column_dimensions[columna[0].column_letter].width = ancho
    #sheet.column_dimensions[columna[0].column_letter].fill = fill
#Guardar el archivo
wb.save(Path(PathC) / 'Outputs' / NameItem / nombre_archivo)


